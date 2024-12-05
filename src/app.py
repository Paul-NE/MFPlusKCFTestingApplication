from pathlib import Path
import logging
import sys
import os
parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if parent_dir not in sys.path:
    sys.path.append(parent_dir)

from openpyxl import Workbook, load_workbook
import pandas as pd

from app_gui.object_selector import get_number
from analytic_tools.nullable_annotation import NullableAnnotation
from analytic_tools.annotation_light import AnnotationLight
from analytic_tools.annotation_json import AnnotationJson, NoSuchObjectFound
from analytic_tools.ious import IOUs, Analytics
from marks.marks import *
import config

from trackers.scale_estimators import estimate_by_affine, scale_by_all_distances
from trackers.forward_backward_pnt_filter import ForwardBackwardPntFilter
from trackers.forward_bachkward_flow import ForwardBachkwardFlow
from trackers.kcf.corellation_tracker import CorellationTracker
from trackers.median_flow_tracker import MedianFlowTracker
from trackers.mf_motion_only import MFMotionOnly
from trackers.mf_scaler import MFScaler
from trackers.kcf.kcf import *

from generators.smart_grid_pts_generator import SmartGridPtsGenerator
from generators.shi_tomasi_generator import ShiTomasiGenerator
from generators.fast_pts_generator import FastPtsGenerator
from generators.limit_pts_generator import limit_pts_generator

from video_processors.video_test import VideoTest
from video_processors.webm_video_writer import WebmVideoWriter
from video_processors.video_test import VideoTest
from video_processors.process_universal import FrameProcessorUni

from app_gui.utils import load_json


app_logger = logging.getLogger(__name__)

def init_annotation(annotation_path, config_options:dict) -> NullableAnnotation:
    if config_options["annotation"]["process"]:
        try: 
            return NullableAnnotation(annotation_path)
        except OSError as e:
            app_logger.warning(f"No annotation file: {annotation_path}")
            return None
        except Exception as e:
            app_logger.warning(f"Unable to read {annotation_path}")
            app_logger.error(f"Error: {e}")
            return None
    return None

def _init_scale_estimation(config_options:dict):
    functions = {
        "scale_by_all_distances": scale_by_all_distances,
        "estimate_by_affine": estimate_by_affine
    }
    options = config_options["median_flow_methods"]["scale_estimator"]
    selected = [k for k,v in options.items() if v is True]
    if len(selected) > 1:
        app_logger.warning(f"Selected multiple options for scale estimation: {selected}. Using first one")
    try:
        choise = selected[0]
    except IndexError:
        app_logger.error("Selected no option for scale estimation!")
        raise ValueError()
    return functions[choise]

def _init_points_generator(config_options:dict):
    functions = {
        "fast_pts_generator": FastPtsGenerator,
        "smart_grid_pts_generator": SmartGridPtsGenerator,
        "shi_tomasi_generator": ShiTomasiGenerator
    }
    limit = config_options["median_flow_methods"]["points_generator"]["limit_by_ellipse"]
    options = config_options["median_flow_methods"]["points_generator"]["algorithm"]
    selected = [k for k,v in options.items() if v is True]
    if len(selected) > 1:
        app_logger.warning(f"Selected multiple options for points generator: {selected}. Using first one")
    try:
        choise = selected[0]
    except IndexError as e:
        app_logger.error(f"Selected no option for points generator! Error: {e}")
        raise ValueError()
    
    if limit:
        return limit_pts_generator(functions[choise]())
    return functions[choise]()

def _init_median_flow_scaler(config_options:dict) -> MFScaler:
    if not config_options["components"]["scale"]:
        return
    # pts_gener = SmartGridPtsGenerator(config._points_density)
    pts_gener = _init_points_generator(config_options)
    fb_flow_generator = ForwardBachkwardFlow(config._flow_generator)
    fb_filter = ForwardBackwardPntFilter(config._max_forward_backward_error)
    scale_estimator = _init_scale_estimation(config_options)
    
    options =MFScaler.Options(**config_options["median_flow_options"]) 
    scaler = MFScaler(
        pts_gener,
        fb_filter,
        fb_flow_generator,
        scale_estimator,
        options=options,
    )
    return scaler

def init_kcf(config_options: dict):
    debug = KCFDebugParams(
        showFeatures=False, 
        showAlphaf=False, 
        showTmpl=False, 
        saveTrackerParams=False
        )
    flags = KCFFlags(
        hog=False, 
        fixed_window=True, 
        multiscale=False, 
        normalizeShift=False, 
        smoothMotion=False
        )
    train_params = KCFTrainParams(
        tmplsz=64, 
        lambdar=0.0001, 
        padding=2.5, 
        output_sigma_factor=0.125, 
        sigma=0.2, 
        interp_factor=0.001
        )
    hog = KCFHogParams(NUM_SECTOR=9, cell_size=4)
    params = KCFParams(flags=flags, hog=hog, debug=debug, train=train_params)
    return KCFTrackerNormal(params)

def _init_corellation(config_options: dict):
    configs = config_options["corellation_tracker"]
    options = CorellationTracker.Options(
        *configs["options"])
    math_params = CorellationTracker.MathParameters(*configs["math_parameters"])
    return CorellationTracker(
        math_parameters=math_params,
        options=options
        )

def post_main(test_folder_path: Path, config_options: dict):
    video_path = test_folder_path / config._video_name
    video_test_options = VideoTest.Options(process_video_options=True)
    frame_process_options = FrameProcessorUni.Options(
        **config_options["frame_processor_options"]
        )

    annotation_path = test_folder_path / config._annotation_name
    annotation = init_annotation(annotation_path, config_options)
    
    video_writer = WebmVideoWriter(config._test_results_write_path) if config_options["components"]["write_video"] else None
    _analytics_engine = Analytics()
    
    scaler = _init_median_flow_scaler(config_options)
    scaler_windows = scaler.get_debug_windows()
    
    tracker = None
    tracker_windows = []
    if config_options["tracking_option"]["corellation"]:
        tracker = _init_corellation(config_options)
        tracker_windows = tracker.get_debug_windows()
    
    windows = scaler_windows + tracker_windows

    operation = FrameProcessorUni(
        tracker=tracker,
        scaler=scaler,
        annotation=annotation, 
        options=frame_process_options,
        video_writer=video_writer,
        windows_to_show=windows,
        analytics_engine=_analytics_engine
        )
    
    v_test = VideoTest(
        video=video_path, 
        operation=operation,
        options=video_test_options
    )
    
    cap = v_test.capoture
    annot_json_path = test_folder_path / "test.json"
    if config_options["annotation"]["process"] and config_options["annotation"]["json"]:
        try:
            json_options = config_options["json_annotation"]
            annot_json = AnnotationJson(annot_json_path, cap)
            
            if json_options["selection"]["preselected"]:
                obj_number = json_options["selected_obj"]
            elif json_options["selection"]["object_selector"]:
                obj_number = get_number(0, annot_json.objects)
            else:
                raise ValueError('Wrong json option "selection"')
            annot_json.selected_object = obj_number
            
            operation.set_annotation_json(annot_json)
        except FileNotFoundError as e:
            app_logger.warning(f"{e.strerror} {annot_json_path}")
    
    v_test.run()
    
    if video_writer is not None:
        video_writer.release()
    
    return _analytics_engine

def write_or_append_to_excel(file_path, sheet_name, df):
    # Check if the file exists
    if not os.path.exists(file_path):
        # File doesn't exist, create a new file and write the DataFrame
        df.to_excel(file_path, sheet_name=sheet_name, index=False)
        print(f"Created new file '{file_path}' and added data to sheet '{sheet_name}'.")
        return
    
    # File exists, load the workbook
    book = load_workbook(file_path)

    if sheet_name in book.sheetnames:
        # Sheet exists, load the sheet into a DataFrame
        existing_df = pd.read_excel(file_path, sheet_name=sheet_name)

        # Check if the sheet is empty
        if existing_df.empty:
            # Write data with headers
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"Sheet '{sheet_name}' was empty. Added new data with headers.")
        else:
            # Append data without rewriting headers
            updated_df = pd.concat([existing_df, df], ignore_index=True)
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                updated_df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"Appended data to existing sheet '{sheet_name}'.")
    else:
        # Sheet doesn't exist, create it
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"Added new sheet '{sheet_name}' to the file.")

def main(test_folder_path: Path, config_options: dict):
    math_model_name = _init_scale_estimation(config_options).__name__
    points_generator_name = type(_init_points_generator(config_options)).__name__
    
    columns = ["Video", "Obj_id", "IOU", "IOU_dispertion", "Samples", "map30", "map40", "map50", "map75"]
    df = pd.DataFrame()
    
    none_data = {
            "Video": None,
            "Obj_id": None, 
            "IOU": None,
            "IOU_dispertion": None,
            "Mean dS error": None,
            "width_1": None,
            "width_2": None,
            "Succes samples": None,
            "All samples": None,
            "map30": None,
            "map40": None,
            "map50": None,
            "map75": None,
            "math_model": math_model_name,
            "points_generator": points_generator_name
        }
    # Process all annotated objects one by one untill get an error
    annotation_index = 0
    while config_options["json_annotation"]["run_all"]:
        options = config_options.copy()
        options["json_annotation"]["selection"]["object_selector"] = False
        options["json_annotation"]["selection"]["preselected"] = True
        options["json_annotation"]["selected_obj"] = annotation_index
        annotation_index+=1
        try:
            result = post_main(test_folder_path, options)
        except NoSuchObjectFound as e:
            app_logger.info(f"Stopped on annotation number {options["json_annotation"]["selected_obj"]}.")
            break
        
        summary = result.summary()
        if not summary:
            data = none_data.copy()
            data["Video"] = test_folder_path
            data["Obj_id"] = annotation_index
        else:
            data = {
                "Video": test_folder_path,
                "Object id": annotation_index,
                "IOU": summary.iou,
                "IOU_dispertion": summary.iou_dispertion,
                "Mean dS error": summary.ds_mean_error,
                "width_1": summary.width_1,
                "width_2": summary.width_2,
                "Succes samples": summary.succes,
                "All samples": summary.all_samples,
                "map30": summary.map30,
                "map40": summary.map40,
                "map50": summary.map50,
                "map75": summary.map75,
                "math_model": math_model_name,
                "points_generator": points_generator_name
            }
        df = pd.concat([df, pd.DataFrame([data])], ignore_index=True)
        print(df)
    else:
        post_main(test_folder_path, config_options)
    
    if config_options["components"]["write_resulting_table"]:
        write_or_append_to_excel("/home/poul/temp/test.xlsx", "test2", df)


if __name__ == "__main__":
    params = load_json("./config_.json")
    
    main(Path(r"/home/poul/temp/Vids/annotated/New/rotate_street_vid_1"), params)
    main(Path(r"/home/poul/temp/Vids/annotated/New/rotate_street_vid_2"), params)
    main(Path(r"/home/poul/temp/Vids/annotated/New/Street_Vid_4"), params)
    main(Path(r"/home/poul/temp/Vids/annotated/New/Street_Vid_5"), params)
    main(Path(r"/home/poul/temp/Vids/annotated/New/StreetVid_1"), params)
    main(Path(r"/home/poul/temp/Vids/annotated/New/StreetVid_2"), params)
    # input()
    
    os._exit(0)
